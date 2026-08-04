"""Ingest stage: load PDF / image / xlsx sources into Document pages."""

from __future__ import annotations

from abc import ABC, abstractmethod
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from prismdoc.errors import InputTooLargeError, UnreadableDocumentError
from prismdoc.models import Block, Document, Page, Source
from prismdoc.pdf import PdfEngine, default_engine
from prismdoc.registry import register
from prismdoc.stages.base import Context, Stage

_IMAGE_EXTENSIONS: tuple[str, ...] = (
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".bmp",
    ".tif",
    ".tiff",
)

_MIME_TO_EXTENSION: dict[str, str] = {
    "application/pdf": ".pdf",
    "image/png": ".png",
    "image/jpeg": ".jpg",
    "image/gif": ".gif",
    "image/webp": ".webp",
    "image/bmp": ".bmp",
    "image/tiff": ".tiff",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
}


class Loader(ABC):
    """Loads a source file into a list of pages."""

    name: str
    extensions: tuple[str, ...]

    @abstractmethod
    def load(self, source: Source) -> list[Page]:
        """Read ``source`` and return pages (no OCR)."""
        ...


class PdfLoader(Loader):
    """Extract text and layout blocks from a PDF.

    The engine is injectable so a caller can trade the permissive default for
    PyMuPDF's speed — see `prismdoc.pdf`. The default is fixed rather than
    "whatever is installed", so a build's licence does not depend on what else
    happened to be in the environment.
    """

    name = "pdf"
    extensions: tuple[str, ...] = (".pdf",)

    def __init__(self, engine: PdfEngine | None = None) -> None:
        self._engine = engine or default_engine()

    def load(self, source: Source) -> list[Page]:
        return self._engine.load_pages(source.path)


class ImageLoader(Loader):
    """Attach an image reference; OCR is deferred to Parse (T-003)."""

    name = "image"
    extensions: tuple[str, ...] = _IMAGE_EXTENSIONS

    def load(self, source: Source) -> list[Page]:
        path = source.path
        try:
            with Image.open(path) as img:
                img.load()
        except Exception as exc:
            raise UnreadableDocumentError(f"Cannot read {path!r}: {exc}") from exc
        return [Page(index=0, text="", image_ref=path)]


class XlsxLoader(Loader):
    """Turn each worksheet into a page of tab/newline-joined cell text."""

    name = "xlsx"
    extensions: tuple[str, ...] = (".xlsx",)

    def load(self, source: Source) -> list[Page]:
        path = source.path
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise UnreadableDocumentError(f"Cannot read {path!r}: {exc}") from exc
        try:
            pages: list[Page] = []
            for index, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=True):
                    cells = ["" if cell is None else str(cell) for cell in row]
                    rows.append("\t".join(cells))
                text = "\n".join(rows)
                pages.append(
                    Page(
                        index=index,
                        text=text,
                        blocks=[
                            Block(
                                text=sheet_name,
                                meta={"kind": "sheet_name"},
                            )
                        ],
                    )
                )
            return pages
        except Exception as exc:
            raise UnreadableDocumentError(f"Cannot read {path!r}: {exc}") from exc
        finally:
            workbook.close()


class TextLoader(Loader):
    """Load plain text / markdown files as a single page."""

    name = "text"
    extensions: tuple[str, ...] = (".txt", ".md")

    def load(self, source: Source) -> list[Page]:
        return [
            Page(
                index=0,
                text=Path(source.path).read_text(encoding="utf-8", errors="replace"),
            )
        ]


class IngestStage(Stage):
    """Select a loader by extension (or MIME) and populate ``doc.pages``."""

    name = "ingest"

    def __init__(self, loaders: list[Loader] | None = None) -> None:
        self.loaders = loaders or [
            PdfLoader(),
            ImageLoader(),
            XlsxLoader(),
            TextLoader(),
        ]
        self._by_extension: dict[str, Loader] = {}
        for loader in self.loaders:
            for ext in loader.extensions:
                self._by_extension[ext.lower()] = loader

    def run(self, doc: Document, ctx: Context) -> Document:
        loader = self._select_loader(doc.source)
        doc.pages = loader.load(doc.source)
        max_pages = ctx.options.get("max_pages")
        if max_pages is not None and len(doc.pages) > max_pages:
            raise InputTooLargeError(
                f"Document has {len(doc.pages)} pages, which exceeds the limit "
                f"of {max_pages} pages"
            )
        return doc

    def _select_loader(self, source: Source) -> Loader:
        ext = Path(source.path).suffix.lower()
        if not ext and source.mime:
            ext = _MIME_TO_EXTENSION.get(source.mime.lower(), "")
        loader = self._by_extension.get(ext)
        if loader is None:
            supported = ", ".join(sorted(self._by_extension))
            raise ValueError(
                f"Unsupported source extension {ext!r} for path {source.path!r}; "
                f"supported: {supported}"
            )
        return loader


def register_plugins() -> None:
    """Register default loaders and ingest stage in the plugin registry."""
    register("loader.pdf", PdfLoader)
    register("loader.image", ImageLoader)
    register("loader.xlsx", XlsxLoader)
    register("loader.text", TextLoader)
    register("ingest.default", IngestStage)


register_plugins()
